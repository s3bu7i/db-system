import json
import posixpath
import re
import zipfile
from contextlib import contextmanager
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import openpyxl
import xlrd
from openpyxl.utils.cell import range_boundaries

from .db import get_conn


XLSX_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SUPPORTED_SUFFIXES = {*XLSX_SUFFIXES, ".xls"}
BATCH_SIZE = 1000
TITLE_SCAN_ROWS = 15
TITLE_SCAN_COLUMNS = 30


def utc_now() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def is_supported_excel(filename: str) -> bool:
    return Path(filename).suffix.lower() in SUPPORTED_SUFFIXES


def import_excel_file(file_id: int, path: str) -> None:
    workbook_kind = _detect_workbook_kind(path)
    detected_name = detect_workbook_name(path)
    with get_conn() as conn:
        conn.execute(
            "UPDATE files SET status = ?, progress = ?, message = ? WHERE id = ?",
            ("importing", 1, "Import basladi", file_id),
        )
        conn.execute("DELETE FROM sheets WHERE file_id = ?", (file_id,))
        conn.commit()

    try:
        if workbook_kind == "xls":
            _import_xls(file_id, path)
        else:
            _import_xlsx(file_id, path)

        with get_conn() as conn:
            if detected_name:
                conn.execute(
                    """
                    UPDATE files
                       SET original_name = ?,
                           status = ?,
                           progress = ?,
                           message = ?,
                           imported_at = ?
                     WHERE id = ?
                    """,
                    (detected_name, "ready", 100, "Hazirdir", utc_now(), file_id),
                )
            else:
                conn.execute(
                    """
                    UPDATE files
                       SET status = ?, progress = ?, message = ?, imported_at = ?
                     WHERE id = ?
                    """,
                    ("ready", 100, "Hazirdir", utc_now(), file_id),
                )
            conn.commit()
    except Exception as exc:  # noqa: BLE001 - stored for the UI/job status.
        with get_conn() as conn:
            conn.execute(
                "UPDATE files SET status = ?, progress = ?, message = ? WHERE id = ?",
                ("failed", 0, str(exc), file_id),
            )
            conn.commit()


def read_sheet_metadata(path: str, sheet_index: int) -> dict[str, Any]:
    if _detect_workbook_kind(path) == "xls":
        workbook = _open_xls_workbook(path)
        try:
            if sheet_index < 1 or sheet_index > workbook.nsheets:
                return {"merged_cells": []}
            return _xls_sheet_metadata(workbook.sheet_by_index(sheet_index - 1))
        finally:
            workbook.release_resources()

    return _xlsx_sheet_metadata(path).get(sheet_index, {"merged_cells": []})


def detect_workbook_name(path: str) -> str:
    try:
        if _detect_workbook_kind(path) == "xls":
            return _detect_xls_workbook_name(path)
        return _detect_xlsx_workbook_name(path)
    except Exception:  # noqa: BLE001 - title detection must not break imports.
        return ""


def _detect_workbook_kind(path: str) -> str:
    suffix = Path(path).suffix.lower()
    if suffix in XLSX_SUFFIXES:
        return "xlsx"
    if suffix == ".xls" and _has_zip_signature(path):
        return "xlsx"
    return "xls"


def _has_zip_signature(path: str) -> bool:
    try:
        with Path(path).open("rb") as file:
            return file.read(4).startswith(b"PK\x03\x04")
    except OSError:
        return False


@contextmanager
def _open_xlsx_workbook(path: str, *, read_only: bool, data_only: bool):
    file = Path(path).open("rb")
    workbook = None
    try:
        workbook = openpyxl.load_workbook(file, read_only=read_only, data_only=data_only)
        yield workbook
    finally:
        if workbook is not None:
            workbook.close()
        file.close()


def _detect_xlsx_workbook_name(path: str) -> str:
    with _open_xlsx_workbook(path, read_only=True, data_only=True) as workbook:
        if not workbook.worksheets:
            return ""
        worksheet = workbook.worksheets[0]
        rows = []
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(TITLE_SCAN_ROWS, worksheet.max_row or TITLE_SCAN_ROWS),
            max_col=min(TITLE_SCAN_COLUMNS, worksheet.max_column or TITLE_SCAN_COLUMNS),
            values_only=True,
        ):
            rows.append([_normalize_title_value(value) for value in row])
        return _best_workbook_name(rows)


def _detect_xls_workbook_name(path: str) -> str:
    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        if workbook.nsheets < 1:
            return ""
        worksheet = workbook.sheet_by_index(0)
        rows = []
        for row_idx in range(min(TITLE_SCAN_ROWS, worksheet.nrows)):
            values = [
                _normalize_title_value(worksheet.cell_value(row_idx, col_idx))
                for col_idx in range(min(TITLE_SCAN_COLUMNS, worksheet.ncols))
            ]
            rows.append(values)
        return _best_workbook_name(rows)
    finally:
        workbook.release_resources()


def _best_workbook_name(rows: list[list[str]]) -> str:
    candidates: list[tuple[int, int, str]] = []
    for row_index, row in enumerate(rows):
        values = [value for value in row if value]
        if not values:
            continue
        if _looks_like_data_row(values):
            continue
        combined = _clean_workbook_name(" ".join(values))
        if not _is_title_candidate(combined):
            continue
        density_bonus = 30 if len(values) <= 2 else 0
        header_penalty = 60 if len(values) > 3 else 0
        score = len(combined) + density_bonus - header_penalty - (row_index * 3)
        candidates.append((score, -row_index, combined))

    if not candidates:
        return ""
    candidates.sort(reverse=True)
    return candidates[0][2]


def _normalize_title_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return _clean_workbook_name(str(value))


def _clean_workbook_name(value: str) -> str:
    value = re.sub(r"\s+", " ", value.replace("\n", " ")).strip()
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', " ", value)
    value = re.sub(r"\s+", " ", value).strip(" .")
    return value[:180]


def _is_title_candidate(value: str) -> bool:
    if len(value) < 6:
        return False
    normalized = value.lower().replace(".", "").strip()
    unit_words = {"ton", "tons", "thsd tons", "min ton", "faiz", "percent", "%"}
    if normalized in unit_words:
        return False
    if not any(char.isalpha() for char in value):
        return False
    numeric_tokens = len(re.findall(r"\b\d+(?:[.,]\d+)?\b", value))
    alpha_tokens = len(re.findall(r"[^\W\d_]{2,}", value, flags=re.UNICODE))
    if numeric_tokens >= 4 and numeric_tokens > alpha_tokens:
        return False
    return True


def _looks_like_data_row(values: list[str]) -> bool:
    numeric_cells = sum(1 for value in values if _looks_numeric(value))
    text_cells = sum(1 for value in values if value and not _looks_numeric(value))
    return len(values) >= 4 and numeric_cells >= 3 and numeric_cells >= text_cells


def _looks_numeric(value: str) -> bool:
    normalized = value.strip().replace(",", ".")
    if not normalized:
        return False
    try:
        float(normalized)
    except ValueError:
        return False
    return True


def _sheet_heading_text(rows: list[list[Any]]) -> str:
    seen: set[str] = set()
    parts: list[str] = []
    for row in rows:
        for value in row[:TITLE_SCAN_COLUMNS]:
            text = _clean_workbook_name(str(value)) if value not in (None, "") else ""
            if not text:
                continue
            normalized = text.casefold()
            if normalized in seen:
                continue
            seen.add(normalized)
            parts.append(text)
    return " ".join(parts)[:3000]


def _import_xlsx(file_id: int, path: str) -> None:
    sheet_metadata = _xlsx_sheet_metadata(path)
    with _open_xlsx_workbook(path, read_only=True, data_only=False) as workbook:
        total_rows = sum((sheet.max_row or 0) for sheet in workbook.worksheets) or 1
        processed_rows = 0
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            sheet_id = _create_sheet(
                file_id,
                worksheet.title,
                sheet_index,
                sheet_metadata.get(sheet_index, {"merged_cells": []}),
            )
            row_count = 0
            column_count = worksheet.max_column or 0
            heading_rows: list[list[Any]] = []
            batch: list[tuple[int, int, str, str]] = []

            for row_number, row in enumerate(worksheet.iter_rows(values_only=False), start=1):
                processed_rows += 1
                values = [_normalize_value(cell.value) for cell in row]
                styles = _xlsx_row_styles(row)
                values = _trim_trailing_empty(values)
                styles = _trim_row_styles(styles, len(values))
                if not values:
                    continue

                if row_number <= TITLE_SCAN_ROWS:
                    heading_rows.append(values)
                row_count += 1
                column_count = max(column_count, len(values))
                batch.append(
                    (
                        sheet_id,
                        row_number,
                        json.dumps(values, ensure_ascii=False),
                        json.dumps(styles, ensure_ascii=False),
                    )
                )

                if len(batch) >= BATCH_SIZE:
                    _insert_rows(batch)
                    batch.clear()
                    _update_progress(file_id, processed_rows, total_rows, worksheet.title)

            if batch:
                _insert_rows(batch)
            _finish_sheet(sheet_id, row_count, column_count, _sheet_heading_text(heading_rows))
            _update_progress(file_id, processed_rows, total_rows, worksheet.title)

def _import_xls(file_id: int, path: str) -> None:
    workbook = _open_xls_workbook(path)
    total_rows = sum(workbook.sheet_by_index(i).nrows for i in range(workbook.nsheets)) or 1
    processed_rows = 0

    try:
        for sheet_index in range(workbook.nsheets):
            worksheet = workbook.sheet_by_index(sheet_index)
            sheet_id = _create_sheet(file_id, worksheet.name, sheet_index + 1, _xls_sheet_metadata(worksheet))
            row_count = 0
            column_count = worksheet.ncols
            heading_rows: list[list[Any]] = []
            batch: list[tuple[int, int, str, str]] = []

            for row_idx in range(worksheet.nrows):
                processed_rows += 1
                values = [
                    _normalize_xls_cell(worksheet.cell(row_idx, col_idx), workbook.datemode)
                    for col_idx in range(worksheet.ncols)
                ]
                styles = _xls_row_styles(workbook, worksheet, row_idx)
                values = _trim_trailing_empty(values)
                styles = _trim_row_styles(styles, len(values))
                if not values:
                    continue

                if row_idx < TITLE_SCAN_ROWS:
                    heading_rows.append(values)
                row_count += 1
                column_count = max(column_count, len(values))
                batch.append(
                    (
                        sheet_id,
                        row_idx + 1,
                        json.dumps(values, ensure_ascii=False),
                        json.dumps(styles, ensure_ascii=False),
                    )
                )

                if len(batch) >= BATCH_SIZE:
                    _insert_rows(batch)
                    batch.clear()
                    _update_progress(file_id, processed_rows, total_rows, worksheet.name)

            if batch:
                _insert_rows(batch)
            _finish_sheet(sheet_id, row_count, column_count, _sheet_heading_text(heading_rows))
            _update_progress(file_id, processed_rows, total_rows, worksheet.name)
    finally:
        workbook.release_resources()


def _create_sheet(file_id: int, name: str, sheet_index: int, meta: dict[str, Any] | None = None) -> int:
    with get_conn() as conn:
        cursor = conn.execute(
            """
            INSERT INTO sheets (file_id, name, sheet_index, meta_json, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (file_id, name, sheet_index, json.dumps(meta or {}, ensure_ascii=False), utc_now()),
        )
        conn.commit()
        return int(cursor.lastrowid)


def _insert_rows(rows: list[tuple[int, int, str, str]]) -> None:
    with get_conn() as conn:
        conn.executemany(
            """
            INSERT OR REPLACE INTO sheet_rows (sheet_id, row_number, cells_json, styles_json)
            VALUES (?, ?, ?, ?)
            """,
            rows,
        )
        conn.commit()


def _xlsx_row_styles(row: tuple[Any, ...]) -> dict[str, list[int]]:
    bold = [
        index
        for index, cell in enumerate(row)
        if bool(getattr(getattr(cell, "font", None), "bold", False))
    ]
    return {"bold": bold} if bold else {}


def _xls_row_styles(workbook: Any, worksheet: Any, row_idx: int) -> dict[str, list[int]]:
    bold: list[int] = []
    for col_idx in range(worksheet.ncols):
        try:
            xf_index = worksheet.cell_xf_index(row_idx, col_idx)
            font_index = workbook.xf_list[xf_index].font_index
            if bool(workbook.font_list[font_index].bold):
                bold.append(col_idx)
        except (AttributeError, IndexError):
            continue
    return {"bold": bold} if bold else {}


def _trim_row_styles(styles: dict[str, list[int]], cell_count: int) -> dict[str, list[int]]:
    bold = [index for index in styles.get("bold", []) if index < cell_count]
    return {"bold": bold} if bold else {}


def _finish_sheet(sheet_id: int, row_count: int, column_count: int, heading_text: str) -> None:
    with get_conn() as conn:
        conn.execute(
            "UPDATE sheets SET row_count = ?, column_count = ?, heading_text = ? WHERE id = ?",
            (row_count, column_count, heading_text, sheet_id),
        )
        conn.commit()


def _update_progress(file_id: int, processed_rows: int, total_rows: int, sheet_name: str) -> None:
    progress = max(1, min(99, int((processed_rows / total_rows) * 100)))
    with get_conn() as conn:
        conn.execute(
            "UPDATE files SET progress = ?, message = ? WHERE id = ?",
            (progress, f"Import edilir: {sheet_name}", file_id),
        )
        conn.commit()


def _normalize_value(value: Any) -> str | int | float | bool:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value)


def _normalize_xls_cell(cell: Any, datemode: int) -> str | int | float | bool:
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, datemode).isoformat()
        except Exception:  # noqa: BLE001 - corrupted dates fall back to raw value.
            return str(cell.value)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return int(cell.value) if float(cell.value).is_integer() else cell.value
    return str(cell.value)


def _trim_trailing_empty(values: list[Any]) -> list[Any]:
    last = len(values)
    while last > 0 and values[last - 1] == "":
        last -= 1
    return values[:last]


def _open_xls_workbook(path: str) -> xlrd.book.Book:
    try:
        return xlrd.open_workbook(path, formatting_info=True, on_demand=True)
    except NotImplementedError:
        return xlrd.open_workbook(path, on_demand=True)


def _xls_sheet_metadata(worksheet: Any) -> dict[str, Any]:
    merged_cells = []
    for row_start, row_end, col_start, col_end in getattr(worksheet, "merged_cells", []) or []:
        if row_end <= row_start or col_end <= col_start:
            continue
        merged_cells.append(
            {
                "r1": row_start + 1,
                "r2": row_end,
                "c1": col_start + 1,
                "c2": col_end,
            }
        )
    return {"merged_cells": merged_cells}


def _xlsx_sheet_metadata(path: str) -> dict[int, dict[str, Any]]:
    try:
        with zipfile.ZipFile(path) as archive:
            relationships = _xlsx_workbook_relationships(archive)
            sheet_paths = _xlsx_sheet_paths(archive, relationships)
            return {
                sheet_index: {"merged_cells": _xlsx_merged_cells(archive, sheet_path)}
                for sheet_index, sheet_path in sheet_paths.items()
            }
    except (KeyError, OSError, zipfile.BadZipFile, ElementTree.ParseError, ValueError):
        return {}


def _xlsx_workbook_relationships(archive: zipfile.ZipFile) -> dict[str, str]:
    root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relationships: dict[str, str] = {}
    for element in root.iter():
        if not element.tag.endswith("Relationship"):
            continue
        rel_id = element.attrib.get("Id")
        target = element.attrib.get("Target")
        if rel_id and target:
            relationships[rel_id] = _xlsx_normalize_target(target)
    return relationships


def _xlsx_sheet_paths(archive: zipfile.ZipFile, relationships: dict[str, str]) -> dict[int, str]:
    root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    sheet_paths: dict[int, str] = {}
    sheet_index = 0
    for element in root.iter():
        if not element.tag.endswith("sheet"):
            continue
        sheet_index += 1
        rel_id = element.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if not rel_id or rel_id not in relationships:
            continue
        sheet_paths[sheet_index] = relationships[rel_id]
    return sheet_paths


def _xlsx_normalize_target(target: str) -> str:
    if target.startswith("/"):
        normalized = target.lstrip("/")
    elif target.startswith("xl/"):
        normalized = target
    else:
        normalized = posixpath.normpath(posixpath.join("xl", target))
    return normalized.replace("\\", "/")


def _xlsx_merged_cells(archive: zipfile.ZipFile, sheet_path: str) -> list[dict[str, int]]:
    root = ElementTree.fromstring(archive.read(sheet_path))
    merged_cells: list[dict[str, int]] = []
    for element in root.iter():
        if not element.tag.endswith("mergeCell"):
            continue
        ref = element.attrib.get("ref")
        if not ref:
            continue
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        if min_row == max_row and min_col == max_col:
            continue
        merged_cells.append({"r1": min_row, "r2": max_row, "c1": min_col, "c2": max_col})
    return merged_cells
